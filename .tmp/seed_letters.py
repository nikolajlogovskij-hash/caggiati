"""Extract letters from Excel and generate SQL seed data."""
import openpyxl
import json

wb = openpyxl.load_workbook("doc/CAGGIATI СКЛАД БУКВЫ+СИМВОЛЫ_1 01.04.xlsm", data_only=True)
ws = wb["Сirilico Romano"]

# Structure: rows 3-43 have data, columns grouped in sets of ~8
# Each letter block: articul | char | ? | ? | ? | ? | ? | ?

letters = []
# We'll extract from the specific data rows
# Looking at the output, data starts at row 4 (after header row 3)
# Columns: A-H for uppercase 5cm, I-P for lowercase 3cm, Q-Y for special chars

# Read the raw data from specific known positions
# Based on the Excel output, let's find data rows
data_found = []

for row_idx in range(3, 43):  # rows 3-42
    cells = []
    for col_idx in range(1, 30):  # scan first 30 columns
        cell = ws.cell(row=row_idx, column=col_idx)
        val = cell.value
        if val is not None:
            cells.append((col_idx, str(val).strip()))
    
    if cells:
        # Look for patterns: articul (00950/XX or 00930/XX) followed by a single char
        for i, (col, val) in enumerate(cells):
            if val.startswith(('00950/', '00930/', '46303/', '46405/', '46507/', '46051/', '46057/')):
                # This is an articul - look for the next value which should be a letter
                if i + 1 < len(cells):
                    char = cells[i + 1][1]
                    if len(char) == 1 and char.isalpha() or char in '.,-/0123456789':
                        data_found.append({
                            'row': row_idx,
                            'col': col,
                            'articul': val,
                            'char': char
                        })

print(f"Found {len(data_found)} letter entries:")
for item in data_found:
    print(f"  {item['articul']} -> '{item['char']}' (row {item['row']}, col {item['col']})")

# Now let's try to get stock and pricing info
# The structure seems to be: articul, char, stock_5cm, ?, ?, price, ?
# Let's read more context around each entry
print("\n\nDetailed context around each entry:")
for item in data_found[:10]:
    row_idx = item['row']
    col_idx = item['col']
    print(f"\n--- {item['articul']} = '{item['char']}' (row {row_idx}, starting col {col_idx}) ---")
    for c in range(col_idx, min(col_idx + 12, 50)):
        val = ws.cell(row=row_idx, column=c).value
        if val is not None:
            print(f"  col {c}: {val}")