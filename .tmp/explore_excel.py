"""Explore the Excel file with letter/symbol inventory."""
import openpyxl

wb = openpyxl.load_workbook("doc/Остатки склада Н.xlsx", data_only=True)
print(f"Sheets: {wb.sheetnames}")

for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"\n=== Sheet: {sn} ===")
    print(f"Dimensions: {ws.dimensions}, rows={ws.max_row}, cols={ws.max_column}")
    # Print first 30 rows
    for row in ws.iter_rows(min_row=1, max_row=min(30, ws.max_row), values_only=False):
        vals = [str(c.value)[:40] if c.value is not None else "" for c in row]
        print(" | ".join(vals))