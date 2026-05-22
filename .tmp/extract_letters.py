"""Extract all letters from Excel and generate SQL seed data."""
import openpyxl
import json

wb = openpyxl.load_workbook("doc/CAGGIATI СКЛАД БУКВЫ+СИМВОЛЫ_1 01.04.xlsm", data_only=True)
ws = wb["Сirilico Romano"]

# Scan all data rows first to understand the structure
print("=== SCANNING ALL ARICULS ===")
found = []
for row_idx in range(1, 80):
    for col_idx in range(1, 50):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val is not None:
            val_str = str(val).strip()
            if '/' in val_str:
                found.append({'row': row_idx, 'col': col_idx, 'val': val_str})

prefixes = {}
for f in found:
    prefix = f['val'].split('/')[0]
    if prefix not in prefixes:
        prefixes[prefix] = []
    prefixes[prefix].append(f)

print(f"Found {len(found)} articul-like values across {len(prefixes)} prefixes")
for p, items in sorted(prefixes.items()):
    print(f"  {p}: {len(items)} items (e.g. '{items[0]['val']}' at row {items[0]['row']}, col {items[0]['col']})")

# Now catalog all unique articuls with their data
# Structure per group (5 cols): articul | char | stock_5cm | stock_3cm | total
# Groups: 1-5, 6-10, 11-15, 16-20, 21-25, 26-30, 31-35, 36-40

valid_prefixes = set(prefixes.keys())
products = []

# Process groups systematically
for row_idx in range(3, 80):
    for group_start in range(1, 46, 5):
        articul_val = ws.cell(row=row_idx, column=group_start).value
        if articul_val is None:
            continue
        articul = str(articul_val).strip()
        if '/' not in articul:
            continue
        
        prefix = articul.split('/')[0]
        if prefix not in valid_prefixes:
            continue
        
        char = ws.cell(row=row_idx, column=group_start + 1).value
        char = str(char).strip() if char is not None else ""
        
        s5 = ws.cell(row=row_idx, column=group_start + 2).value
        s3 = ws.cell(row=row_idx, column=group_start + 3).value
        total = ws.cell(row=row_idx, column=group_start + 4).value
        
        s5 = int(s5) if s5 is not None else 0
        s3 = int(s3) if s3 is not None else 0
        total = int(total) if total is not None else 0
        
        products.append({
            'articul': articul,
            'char': char,
            'stock_5cm': s5,
            'stock_3cm': s3,
            'total': total,
            'row': row_idx,
            'col': group_start
        })

print(f"\n=== EXTRACTED {len(products)} PRODUCTS ===")

# Determine category for each prefix
prefix_categories = {
    '00950': {'height': '5cm', 'language': 'Cyrillic', 'case': 'uppercase', 'category': 'letter'},
    '00930': {'height': '3cm', 'language': 'Cyrillic', 'case': 'lowercase', 'category': 'letter'},
    '00940': {'height': '3cm', 'language': 'Latin', 'case': 'lowercase', 'category': 'letter'},
    '00960': {'height': '5cm', 'language': 'Cyrillic', 'case': 'uppercase', 'category': 'letter'},
    '00965': {'height': '5cm', 'language': 'Latin', 'case': 'uppercase', 'category': 'letter'},
    '00970': {'height': '3cm', 'language': 'Cyrillic', 'case': 'lowercase', 'category': 'letter'},
    '00975': {'height': '3cm', 'language': 'Latin', 'case': 'lowercase', 'category': 'letter'},
    '46010': {'height': '3cm', 'language': None, 'case': 'number', 'category': 'number'},
    '46015': {'height': '5cm', 'language': None, 'case': 'number', 'category': 'number'},
    '46052': {'height': '3cm', 'language': None, 'case': 'number', 'category': 'number'},
    '46058': {'height': '3cm', 'language': None, 'case': 'number', 'category': 'number'},
    '46303': {'height': '3cm', 'language': None, 'case': 'number', 'category': 'number'},
    '46305': {'height': '5cm', 'language': None, 'case': 'number', 'category': 'number'},
    '46405': {'height': '5cm', 'language': None, 'case': 'number', 'category': 'number'},
    '46507': {'height': '3cm', 'language': None, 'case': 'symbol', 'category': 'symbol'},
    '46515': {'height': '5cm', 'language': None, 'case': 'symbol', 'category': 'symbol'},
    '46516': {'height': '5cm', 'language': None, 'case': 'symbol', 'category': 'symbol'},
    '46051': {'height': '3cm', 'language': None, 'case': 'number', 'category': 'number'},
    '46057': {'height': '3cm', 'language': None, 'case': 'number', 'category': 'number'},
}

for p in products:
    prefix = p['articul'].split('/')[0]
    cat = prefix_categories.get(prefix, {})
    p['height'] = cat.get('height', 'unknown')
    p['language'] = cat.get('language')
    p['letter_case'] = cat.get('case')
    p['category'] = cat.get('category', 'other')

# Print summary grouped by prefix
print("\n=== SUMMARY BY PREFIX ===")
for prefix in sorted(prefix_categories.keys()):
    items = [p for p in products if p['articul'].startswith(prefix)]
    if items:
        cat = prefix_categories[prefix]
        print(f"\n{prefix} ({cat['category']}, {cat['height']}, {cat.get('case', 'N/A')}): {len(items)} items")
        total_s5 = sum(p['stock_5cm'] for p in items)
        total_s3 = sum(p['stock_3cm'] for p in items)
        print(f"  Total stock: {total_s5} (5cm) / {total_s3} (3cm)")

# Print detailed list
print("\n=== DETAILED LIST ===")
for p in products:
    print(f"  {p['articul']:12s} '{p['char']:3s}'  5cm:{p['stock_5cm']:4d}  3cm:{p['stock_3cm']:4d}  total:{p['total']:4d}  [{p['height']} {p['language'] or ''} {p['letter_case']}]")

# Generate SQL
print("\n\n=== SQL INSERT ===")
# First check existing table structure
print("-- First, let's add columns if needed:")
print("-- ALTER TABLE letters ADD COLUMN IF NOT EXISTS articul TEXT UNIQUE;")
print("-- ALTER TABLE letters ADD COLUMN IF NOT EXISTS letter TEXT;")
print("-- ALTER TABLE letters ADD COLUMN IF NOT EXISTS height TEXT;")
print("-- ALTER TABLE letters ADD COLUMN IF NOT EXISTS category TEXT;")
print("-- ALTER TABLE letters ADD COLUMN IF NOT EXISTS stock_5cm INTEGER DEFAULT 0;")
print("-- ALTER TABLE letters ADD COLUMN IF NOT EXISTS stock_3cm INTEGER DEFAULT 0;")
print("-- ALTER TABLE letters ADD COLUMN IF NOT EXISTS total_stock INTEGER DEFAULT 0;")
print()

print("INSERT INTO letters (articul, letter, height, category, stock_5cm, stock_3cm, total_stock) VALUES")
values = []
for p in products:
    height = f"'{p['height']}'" if p['height'] else 'NULL'
    values.append(
        f"  ('{p['articul']}', '{p['char']}', {height}, '{p['category']}', "
        f"{p['stock_5cm']}, {p['stock_3cm']}, {p['total']})"
    )
print(",\n".join(values) + ";")

# Save JSON
with open('.tmp/letters_data.json', 'w', encoding='utf-8') as f:
    json.dump(products, f, ensure_ascii=False, indent=2)

print(f"\nSaved {len(products)} products to .tmp/letters_data.json")